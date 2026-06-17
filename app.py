"""
校园智能广播定时系统 V4.0 - Flask Web应用
全Web架构，7个页面全部API
"""

import os
import sys
import json
import logging
import hashlib
from datetime import datetime, date, timedelta
from pathlib import Path
from io import BytesIO

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, make_response
from models import get_db, init_db, add_log, get_setting, set_setting
from engine import engine

os.makedirs("logs", exist_ok=True)
os.makedirs("static/sounds", exist_ok=True)
os.makedirs("data", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/school_bell.log", encoding="utf-8"),
    ]
)
logger = logging.getLogger("SchoolBell")

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB上传限制
app.config['SECRET_KEY'] = 'school-bell-v4-secret'


# ====== 简单密码保护 ======
def check_auth():
    """检查是否需要密码认证"""
    pwd = get_setting("password", "")
    if not pwd:
        return True  # 未设置密码，无需认证
    token = request.headers.get("X-Auth-Token", "") or request.cookies.get("auth_token", "")
    return token == pwd


@app.before_request
def auth_middleware():
    """认证中间件"""
    # 静态文件和登录接口不需要认证
    if request.path.startswith("/static") or request.path == "/api/login":
        return None
    pwd = get_setting("password", "")
    if pwd and not check_auth():
        if request.path.startswith("/api/"):
            return jsonify({"error": "未认证", "need_auth": True}), 401
        return render_template("login.html")


# ====== 页面路由 ======
@app.route("/")
def index():
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# ====== API: 认证 ======
@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(silent=True) or {}
    pwd = get_setting("password", "")
    input_pwd = data.get("password", "")
    if not pwd:
        return jsonify({"success": True, "message": "未设置密码"})
    if hashlib.md5(input_pwd.encode()).hexdigest() == pwd:
        resp = jsonify({"success": True})
        resp.set_cookie("auth_token", pwd, max_age=86400 * 30)
        return resp
    return jsonify({"success": False, "message": "密码错误"}), 403


# ====== API: 系统状态 ======
@app.route("/api/status")
def api_status():
    conn = get_db()
    active = conn.execute("SELECT * FROM schedules WHERE is_active=1").fetchone()
    zones = conn.execute("SELECT * FROM zones ORDER BY sort_order").fetchall()
    schedules = conn.execute("SELECT id, name, is_active FROM schedules ORDER BY id").fetchall()

    # 今日和明日状态
    today_str = date.today().isoformat()
    tomorrow_str = (date.today() + timedelta(days=1)).isoformat()

    zone_list = []
    for z in zones:
        # 今日状态
        today_status = engine.get_zone_status(z["id"], today_str)
        # 明日状态
        tomorrow_override = conn.execute(
            "SELECT * FROM tomorrow_overrides WHERE date=? AND zone_id=?",
            (tomorrow_str, z["id"])
        ).fetchone()

        zone_list.append({
            "id": z["id"],
            "name": z["name"],
            "enabled": bool(z["enabled"]),
            "paused": engine.player.is_paused(z["id"]),
            "playing": engine.player.is_playing(z["id"]),
            "volume": round(engine.player.zone_volumes.get(z["id"], z["volume"]) * 100),
            "device": z["audio_device"] or "默认设备",
            "schedule_id": z["schedule_id"] if "schedule_id" in z.keys() else 0,
            "today": today_status,
            "tomorrow_override": tomorrow_override["action"] if tomorrow_override else None,
            "tomorrow_schedule_id": tomorrow_override["schedule_id"] if tomorrow_override and "schedule_id" in tomorrow_override.keys() else 0,
        })

    next_bells = engine.get_next_bell()

    # NTP状态
    ntp_last_sync = get_setting("ntp_last_sync", "")
    ntp_server = get_setting("ntp_server", "ntp.aliyun.com")
    ntp_offset = engine.ntp.offset if engine.ntp.last_sync else None

    conn.close()
    return jsonify({
        "success": True,
        "current_time": datetime.now().strftime("%H:%M:%S"),
        "current_date": date.today().isoformat(),
        "weekday": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][date.today().weekday()],
        "active_schedule": active["name"] if active else "无",
        "active_schedule_id": active["id"] if active else 0,
        "should_bell_today": engine.holiday.should_bell_today(),
        "global_paused": engine.player.is_global_paused(),
        "zones": zone_list,
        "next_bells": next_bells,
        "ntp": {
            "last_sync": ntp_last_sync or "从未",
            "server": ntp_server,
            "offset": round(ntp_offset, 3) if ntp_offset is not None else None,
            "status": "synced" if ntp_last_sync else "never",
        },
        "schedules": [{"id": s["id"], "name": s["name"], "is_active": bool(s["is_active"])} for s in schedules],
    })


# ====== API: 课表管理 ======
@app.route("/api/schedules", methods=["GET"])
def api_schedules_list():
    zone_filter = request.args.get("zone_id")
    conn = get_db()
    schedules = conn.execute("SELECT * FROM schedules ORDER BY id").fetchall()
    result = []
    for s in schedules:
        tasks = conn.execute(
            "SELECT * FROM schedule_tasks WHERE schedule_id=? ORDER BY zone_id, time",
            (s["id"],)
        ).fetchall()
        # 按区域分组
        zone_tasks = {}
        for t in tasks:
            zid = t["zone_id"]
            if zid not in zone_tasks:
                zone_tasks[zid] = []
            zone_tasks[zid].append(dict(t))
        # 如果指定了区域过滤，只返回包含该区域任务的课表
        if zone_filter and zone_filter not in zone_tasks:
            continue
        result.append({
            "id": s["id"],
            "name": s["name"],
            "is_active": bool(s["is_active"]),
            "zone_tasks": zone_tasks,
            "task_count": len(tasks),
        })
    conn.close()
    return jsonify({"success": True, "schedules": result})


@app.route("/api/schedules", methods=["POST"])
def api_schedule_create():
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "课表名称不能为空"}), 400

    conn = get_db()
    try:
        conn.execute("INSERT INTO schedules (name, is_active) VALUES (?,0)", (name,))
        conn.commit()
        add_log("INFO", "schedule", f"创建课表: {name}")
        return jsonify({"success": True, "message": f"课表'{name}'已创建"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/schedules/<int:sid>", methods=["PUT"])
def api_schedule_update(sid):
    data = request.get_json(silent=True) or {}
    conn = get_db()

    if "name" in data:
        conn.execute("UPDATE schedules SET name=? WHERE id=?", (data["name"], sid))
    if data.get("activate"):
        conn.execute("UPDATE schedules SET is_active=0")
        conn.execute("UPDATE schedules SET is_active=1 WHERE id=?", (sid,))

    conn.commit()
    conn.close()
    add_log("INFO", "schedule", f"更新课表ID={sid}")
    return jsonify({"success": True})


@app.route("/api/schedules/<int:sid>", methods=["DELETE"])
def api_schedule_delete(sid):
    conn = get_db()
    active = conn.execute("SELECT is_active FROM schedules WHERE id=?", (sid,)).fetchone()
    if active and active["is_active"]:
        conn.close()
        return jsonify({"error": "不能删除激活中的课表"}), 400
    conn.execute("DELETE FROM schedule_tasks WHERE schedule_id=?", (sid,))
    conn.execute("DELETE FROM schedules WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    add_log("INFO", "schedule", f"删除课表ID={sid}")
    return jsonify({"success": True})


# ====== API: 课表任务管理 ======
@app.route("/api/tasks", methods=["POST"])
def api_task_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO schedule_tasks (schedule_id, zone_id, time, bell_type, task_name, days, one_time_date, bell_id) VALUES (?,?,?,?,?,?,?,?)",
            (
                data.get("schedule_id"),
                data.get("zone_id", "A"),
                data.get("time", "08:00"),
                data.get("bell_type", "class_start"),
                data.get("task_name", ""),
                data.get("days", "1,2,3,4,5"),
                data.get("one_time_date", ""),
                data.get("bell_id", 0),
            )
        )
        conn.commit()
        add_log("INFO", "task", f"添加打铃任务: {data.get('task_name','')}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/tasks/<int:tid>", methods=["GET"])
def api_task_get(tid):
    """获取单个任务详情"""
    conn = get_db()
    task = conn.execute(
        "SELECT * FROM schedule_tasks WHERE id=?",
        (tid,)
    ).fetchone()
    conn.close()
    
    if task:
        return jsonify({"success": True, "task": dict(task)})
    else:
        return jsonify({"error": "任务不存在"}), 404


@app.route("/api/tasks/<int:tid>", methods=["PUT"])
def api_task_update(tid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    fields = []
    values = []
    for key in ["time", "bell_type", "task_name", "days", "zone_id", "one_time_date", "bell_id"]:
        if key in data:
            fields.append(f"{key}=?")
            values.append(data[key])
    if fields:
        values.append(tid)
        conn.execute(f"UPDATE schedule_tasks SET {','.join(fields)} WHERE id=?", values)
        conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/tasks/<int:tid>", methods=["DELETE"])
def api_task_delete(tid):
    conn = get_db()
    conn.execute("DELETE FROM schedule_tasks WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ====== API: 区域管理 ======
@app.route("/api/zones", methods=["GET"])
def api_zones_list():
    conn = get_db()
    zones = conn.execute("SELECT * FROM zones ORDER BY sort_order").fetchall()
    result = []
    for z in zones:
        result.append({
            "id": z["id"],
            "name": z["name"],
            "enabled": bool(z["enabled"]),
            "audio_device": z["audio_device"],
            "volume": round(engine.player.zone_volumes.get(z["id"], z["volume"]) * 100),
            "playing": engine.player.is_playing(z["id"]),
            "paused": engine.player.is_paused(z["id"]),
            "schedule_id": z["schedule_id"] if "schedule_id" in z.keys() else 0,
        })
    conn.close()
    return jsonify({"success": True, "zones": result})


@app.route("/api/zones", methods=["POST"])
def api_zone_create():
    data = request.get_json(silent=True) or {}
    zone_id = data.get("id", "").strip().upper()
    name = data.get("name", "").strip()
    audio_device = data.get("audio_device", "default")
    volume = float(data.get("volume", 80)) / 100.0
    if not zone_id:
        return jsonify({"error": "区域ID不能为空"}), 400
    conn = get_db()
    existing = conn.execute("SELECT id FROM zones WHERE id=?", (zone_id,)).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": f"区域 {zone_id} 已存在"}), 400
    engine.player.zone_devices[zone_id] = audio_device
    conn.execute(
        "INSERT INTO zones (id, name, enabled, audio_device, volume, sort_order) VALUES (?,?,1,?,?,?)",
        (zone_id, name, audio_device, volume, 999)
    )
    conn.commit()
    conn.close()
    add_log("INFO", "zone", f"创建区域: {zone_id} {name}")
    return jsonify({"success": True, "zone_id": zone_id})


@app.route("/api/zones/<zone_id>", methods=["PUT"])
def api_zone_update(zone_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    fields = []
    values = []
    for key in ["name", "audio_device", "enabled", "schedule_id"]:
        if key in data:
            fields.append(f"{key}=?")
            val = data[key]
            if key == "enabled":
                val = 1 if val else 0
            if key == "audio_device":
                engine.player.zone_devices[zone_id] = val
            values.append(val)
    if "volume" in data:
        vol = float(data["volume"]) / 100.0
        engine.player.zone_volumes[zone_id] = vol
        fields.append("volume=?")
        values.append(vol)
    if fields:
        values.append(zone_id)
        conn.execute(f"UPDATE zones SET {','.join(fields)} WHERE id=?", values)
        conn.commit()
    conn.close()
    add_log("INFO", "zone", f"更新区域设置: {zone_id}")
    return jsonify({"success": True})


@app.route("/api/zones/<zone_id>", methods=["DELETE"])
def api_zone_delete(zone_id):
    conn = get_db()
    conn.execute("DELETE FROM bell_bindings WHERE zone_id=?", (zone_id,))
    conn.execute("DELETE FROM schedule_tasks WHERE zone_id=?", (zone_id,))
    conn.execute("DELETE FROM today_exceptions WHERE zone_id=?", (zone_id,))
    conn.execute("DELETE FROM tomorrow_overrides WHERE zone_id=?", (zone_id,))
    conn.execute("DELETE FROM bell_schedules WHERE zone_id=?", (zone_id,))
    conn.execute("DELETE FROM zones WHERE id=?", (zone_id,))
    conn.commit()
    conn.close()
    add_log("INFO", "zone", f"删除区域: {zone_id}")
    return jsonify({"success": True})


# ====== API: 铃声管理 ======
@app.route("/api/bells", methods=["GET"])
def api_bells_list():
    conn = get_db()
    bells = conn.execute("SELECT * FROM bells ORDER BY bell_type, name").fetchall()
    bindings = conn.execute("SELECT * FROM bell_bindings").fetchall()

    binding_map = {}
    for b in bindings:
        key = f"{b['zone_id']}_{b['bell_type']}"
        binding_map[key] = b["bell_id"]

    result = []
    for b in bells:
        result.append({
            "id": b["id"],
            "name": b["name"],
            "filename": b["filename"],
            "bell_type": b["bell_type"],
            "duration": b["duration"],
            "uploaded_at": b["uploaded_at"],
        })
    conn.close()
    return jsonify({"success": True, "bells": result, "bindings": binding_map})


@app.route("/api/bells/upload", methods=["POST"])
def api_bell_upload():
    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "文件名为空"}), 400

    # 保存文件
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".mp3", ".wav", ".ogg", ".m4a", ".aac", ".flac", ".wma"):
        return jsonify({"error": f"不支持的格式{ext}，支持: MP3/WAV/OGG/M4A/AAC/FLAC/WMA"}), 400

    safe_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{f.filename}"
    filepath = os.path.join("static", "sounds", safe_name)
    f.save(filepath)

    # 自动将MP3转为WAV（避免依赖ffmpeg）
    if ext == ".mp3":
        try:
            import pygame, soundfile as sf, numpy as np
            if not pygame.mixer.get_init():
                pygame.mixer.init(frequency=44100, size=-16, channels=2, buffer=2048)
            snd = pygame.mixer.Sound(filepath)
            arr = pygame.sndarray.samples(snd)
            if arr.dtype != np.float32:
                arr = arr.astype(np.float32) / 32768.0
            sr = pygame.mixer.get_init()[0]
            wav_name = safe_name.replace(".mp3", ".wav")
            wav_path = os.path.join("static", "sounds", wav_name)
            sf.write(wav_path, arr, sr)
            os.remove(filepath)
            safe_name = wav_name
            logger.info(f"MP3已自动转换为WAV: {wav_name}")
        except Exception as e:
            logger.warning(f"MP3转WAV失败（将使用原始MP3）: {e}")

    # 写入数据库
    bell_name = request.form.get("name", os.path.splitext(f.filename)[0])
    bell_type = request.form.get("bell_type", "custom")

    conn = get_db()
    conn.execute(
        "INSERT INTO bells (name, filename, bell_type, duration, uploaded_at) VALUES (?,?,?,?,?)",
        (bell_name, safe_name, bell_type, 0, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()

    add_log("INFO", "bell", f"上传铃声: {bell_name}")
    return jsonify({"success": True, "message": f"铃声'{bell_name}'已上传"})


@app.route("/api/bells/<int:bell_id>", methods=["DELETE"])
def api_bell_delete(bell_id):
    conn = get_db()
    bell = conn.execute("SELECT * FROM bells WHERE id=?", (bell_id,)).fetchone()
    if bell:
        filepath = os.path.join("static", "sounds", bell["filename"])
        if os.path.exists(filepath):
            os.remove(filepath)
        conn.execute("DELETE FROM bell_bindings WHERE bell_id=?", (bell_id,))
        conn.execute("DELETE FROM bells WHERE id=?", (bell_id,))
        conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/bell_bindings", methods=["POST"])
def api_bell_bindings_update():
    """更新铃声类型绑定"""
    data = request.get_json(silent=True) or {}
    zone_id = data.get("zone_id")
    bell_type = data.get("bell_type")
    bell_id = data.get("bell_id")

    conn = get_db()
    conn.execute(
        "DELETE FROM bell_bindings WHERE zone_id=? AND bell_type=?",
        (zone_id, bell_type)
    )
    if bell_id:
        conn.execute(
            "INSERT INTO bell_bindings (zone_id, bell_type, bell_id) VALUES (?,?,?)",
            (zone_id, bell_type, bell_id)
        )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ====== API: 打铃控制 ======
@app.route("/api/bell/ring", methods=["POST"])
def api_bell_ring():
    """手动打铃"""
    data = request.get_json(silent=True) or {}
    zone_id = data.get("zone_id", "A")
    bell_type = data.get("bell_type", "class_start")
    result = engine.manual_bell(zone_id, bell_type)
    return jsonify({"success": result, "message": f"手动打铃: 区域{zone_id} {bell_type}"})


@app.route("/api/bell/pause", methods=["POST"])
def api_bell_pause():
    """暂停"""
    data = request.get_json(silent=True) or {}
    zone_id = data.get("zone_id")
    if zone_id:
        engine.player.pause_zone(zone_id)
        return jsonify({"success": True, "message": f"区域{zone_id}已暂停"})
    else:
        engine.player.pause_global()
        return jsonify({"success": True, "message": "全局已暂停"})


@app.route("/api/bell/resume", methods=["POST"])
def api_bell_resume():
    """恢复"""
    data = request.get_json(silent=True) or {}
    zone_id = data.get("zone_id")
    if zone_id:
        engine.player.resume_zone(zone_id)
        return jsonify({"success": True, "message": f"区域{zone_id}已恢复"})
    else:
        engine.player.resume_global()
        return jsonify({"success": True, "message": "全局已恢复"})


@app.route("/api/bell/stop", methods=["POST"])
def api_bell_stop():
    """停止播放"""
    data = request.get_json(silent=True) or {}
    zone_id = data.get("zone_id")
    if zone_id:
        engine.player.stop_zone(zone_id)
        return jsonify({"success": True, "message": f"区域{zone_id}已停止"})
    else:
        engine.player.stop_all()
        return jsonify({"success": True, "message": "所有播放已停止"})


# ====== API: 今日控制（按区域） ======
@app.route("/api/today_control", methods=["POST"])
def api_today_control():
    """今日打铃控制（按区域）：ring/no_ring/auto"""
    data = request.get_json(silent=True) or {}
    action = data.get("action", "")  # ring=今日打铃, no_ring=今日不打铃, auto=恢复自动
    zone_id = data.get("zone_id", "")  # 区域ID（必填）
    schedule_id = data.get("schedule_id", 0)  # 可选：指定课表

    if not zone_id:
        return jsonify({"error": "请指定区域(zone_id)"}), 400

    today_str = date.today().isoformat()
    conn = get_db()

    # 验证区域存在
    zone = conn.execute("SELECT * FROM zones WHERE id=?", (zone_id,)).fetchone()
    if not zone:
        conn.close()
        return jsonify({"error": "区域不存在"}), 400

    if action == "auto":
        # 恢复自动：删除该区域今天的例外
        conn.execute("DELETE FROM today_exceptions WHERE date=? AND zone_id=?", (today_str, zone_id))
        conn.commit()
        conn.close()
        add_log("INFO", "exception", f"{zone['name']}今日控制：恢复自动")
        return jsonify({"success": True, "message": f"{zone['name']}已恢复自动模式"})

    elif action == "no_ring":
        # 今日不打铃
        conn.execute("DELETE FROM today_exceptions WHERE date=? AND zone_id=?", (today_str, zone_id))
        conn.execute(
            "INSERT INTO today_exceptions (date, zone_id, action, reason, created_at, schedule_id) VALUES (?,?,?,?,?,?)",
            (today_str, zone_id, "force_no_ring", f"{zone['name']}手动设置今日不打铃", datetime.now().isoformat(), 0)
        )
        conn.commit()
        conn.close()
        add_log("INFO", "exception", f"{zone['name']}今日控制：今日不打铃")
        return jsonify({"success": True, "message": f"{zone['name']}已设置今日不打铃"})

    elif action == "ring":
        # 今日打铃
        conn.execute("DELETE FROM today_exceptions WHERE date=? AND zone_id=?", (today_str, zone_id))
        schedule_name = ""
        if schedule_id:
            s = conn.execute("SELECT name FROM schedules WHERE id=?", (schedule_id,)).fetchone()
            schedule_name = s["name"] if s else ""
        conn.execute(
            "INSERT INTO today_exceptions (date, zone_id, action, reason, created_at, schedule_id) VALUES (?,?,?,?,?,?)",
            (today_str, zone_id, "force_ring", f"{zone['name']}手动设置今日打铃" + (f"（{schedule_name}）" if schedule_name else ""), datetime.now().isoformat(), schedule_id or 0)
        )
        conn.commit()
        conn.close()
        add_log("INFO", "exception", f"{zone['name']}今日控制：今日打铃（课表: {schedule_name or '默认'}）")
        return jsonify({"success": True, "message": f"{zone['name']}已设置今日打铃" + (f"（{schedule_name}）" if schedule_name else "")})

    else:
        conn.close()
        return jsonify({"error": "无效操作，请使用 ring/no_ring/auto"}), 400


# ====== API: 明日控制（按区域） ======
@app.route("/api/tomorrow_control", methods=["POST"])
def api_tomorrow_control():
    """明日打铃控制（按区域）：ring/no_ring/auto"""
    data = request.get_json(silent=True) or {}
    action = data.get("action", "")
    zone_id = data.get("zone_id", "")
    schedule_id = data.get("schedule_id", 0)

    if not zone_id:
        return jsonify({"error": "请指定区域(zone_id)"}), 400

    tomorrow_str = (date.today() + timedelta(days=1)).isoformat()
    conn = get_db()

    zone = conn.execute("SELECT * FROM zones WHERE id=?", (zone_id,)).fetchone()
    if not zone:
        conn.close()
        return jsonify({"error": "区域不存在"}), 400

    if action == "auto":
        conn.execute("DELETE FROM tomorrow_overrides WHERE date=? AND zone_id=?", (tomorrow_str, zone_id))
        conn.commit()
        conn.close()
        add_log("INFO", "exception", f"{zone['name']}明日控制：恢复自动")
        return jsonify({"success": True, "message": f"{zone['name']}明日已恢复自动模式"})

    elif action == "no_ring":
        conn.execute("DELETE FROM tomorrow_overrides WHERE date=? AND zone_id=?", (tomorrow_str, zone_id))
        conn.execute(
            "INSERT INTO tomorrow_overrides (date, zone_id, action, schedule_id, reason, created_at) VALUES (?,?,?,?,?,?)",
            (tomorrow_str, zone_id, "force_no_ring", 0, f"{zone['name']}设置明日不打铃", datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        add_log("INFO", "exception", f"{zone['name']}明日控制：明日不打铃")
        return jsonify({"success": True, "message": f"{zone['name']}已设置明日不打铃"})

    elif action == "ring":
        conn.execute("DELETE FROM tomorrow_overrides WHERE date=? AND zone_id=?", (tomorrow_str, zone_id))
        schedule_name = ""
        if schedule_id:
            s = conn.execute("SELECT name FROM schedules WHERE id=?", (schedule_id,)).fetchone()
            schedule_name = s["name"] if s else ""
        conn.execute(
            "INSERT INTO tomorrow_overrides (date, zone_id, action, schedule_id, reason, created_at) VALUES (?,?,?,?,?,?)",
            (tomorrow_str, zone_id, "force_ring", schedule_id or 0, f"{zone['name']}设置明日打铃" + (f"（{schedule_name}）" if schedule_name else ""), datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        add_log("INFO", "exception", f"{zone['name']}明日控制：明日打铃（课表: {schedule_name or '默认'}）")
        return jsonify({"success": True, "message": f"{zone['name']}已设置明日打铃" + (f"（{schedule_name}）" if schedule_name else "")})

    else:
        conn.close()
        return jsonify({"error": "无效操作"}), 400


# ====== API: 今日例外（暂停/响铃/不响） ======
@app.route("/api/today_exceptions", methods=["GET"])
def api_today_exceptions():
    """获取今日例外列表"""
    d = request.args.get("date", date.today().isoformat())
    conn = get_db()
    exceptions = conn.execute(
        "SELECT * FROM today_exceptions WHERE date=?", (d,)
    ).fetchall()
    conn.close()
    return jsonify({"success": True, "exceptions": [dict(e) for e in exceptions]})


@app.route("/api/today_exceptions", methods=["POST"])
def api_today_exception_add():
    """添加今日例外"""
    data = request.get_json(silent=True) or {}
    target_date = data.get("date", date.today().isoformat())
    zone_id = data.get("zone_id", "")
    action = data.get("action", "")  # pause_all, pause_zone, no_bell
    reason = data.get("reason", "")

    if action not in ("pause_all", "pause_zone", "no_bell"):
        return jsonify({"error": "无效操作"}), 400

    conn = get_db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO today_exceptions (date, zone_id, action, reason, created_at) VALUES (?,?,?,?,?)",
            (target_date, zone_id, action, reason, datetime.now().isoformat())
        )
        conn.commit()
        add_log("INFO", "exception", f"今日例外: {action} 区域{zone_id} 原因:{reason}")
        return jsonify({"success": True, "message": f"已设置: {action}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/today_exceptions/<int:eid>", methods=["DELETE"])
def api_today_exception_delete(eid):
    """删除今日例外"""
    conn = get_db()
    conn.execute("DELETE FROM today_exceptions WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ====== API: 节假日 ======
@app.route("/api/holidays", methods=["GET"])
def api_holidays():
    year = int(request.args.get("year", date.today().year))
    engine.holiday.fetch_year(year)

    # 获取自定义节假日
    conn = get_db()
    custom = conn.execute("SELECT * FROM custom_holidays WHERE date LIKE ?", (f"{year}%",)).fetchall()
    conn.close()

    api_list = engine.holiday.get_all_list(year)
    upcoming = engine.holiday.get_upcoming(30)

    # 合并自定义
    for ch in custom:
        api_list.append({
            "date": ch["date"],
            "name": ch["name"],
            "type": "holiday" if ch["is_holiday"] else "workday",
            "custom": True,
            "id": ch["id"],
        })
    api_list.sort(key=lambda x: x["date"])

    return jsonify({
        "success": True,
        "holidays": api_list,
        "upcoming": upcoming,
        "should_bell_today": engine.holiday.should_bell_today(),
    })


@app.route("/api/holidays/sync", methods=["POST"])
def api_holidays_sync():
    """同步节假日数据（强制刷新）"""
    year = date.today().year
    # 强制重新获取：清除缓存
    engine.holiday._fetched_years.discard(year)
    success = engine.holiday.fetch_year(year)
    if success:
        count = len(engine.holiday.holidays)
        add_log("INFO", "holiday", f"同步节假日数据: {year}年 {count}个")
        return jsonify({"success": True, "message": f"已同步{year}年节假日数据，共{count}个"})
    return jsonify({"error": "同步失败，请检查网络连接"}), 500


@app.route("/api/holidays/custom", methods=["POST"])
def api_holiday_custom_add():
    """添加自定义节假日"""
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO custom_holidays (date, name, is_holiday) VALUES (?,?,?)",
            (data.get("date"), data.get("name", ""), 1 if data.get("is_holiday", True) else 0)
        )
        conn.commit()
        add_log("INFO", "holiday", f"自定义假期: {data.get('date')} {data.get('name')}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/holidays/custom/<int:hid>", methods=["PUT"])
def api_holiday_custom_update(hid):
    """更新自定义节假日（切换打铃/不打铃）"""
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        if "is_holiday" in data:
            conn.execute(
                "UPDATE custom_holidays SET is_holiday=? WHERE id=?",
                (1 if data["is_holiday"] else 0, hid)
            )
            conn.commit()
            action = "不打铃(放假)" if data["is_holiday"] else "打铃(调休)"
            add_log("INFO", "holiday", f"更新假期ID={hid}: {action}")
            return jsonify({"success": True, "message": f"已切换为{action}"})
        return jsonify({"error": "缺少is_holiday参数"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/holidays/custom/<int:hid>", methods=["DELETE"])
def api_holiday_custom_delete(hid):
    conn = get_db()
    conn.execute("DELETE FROM custom_holidays WHERE id=?", (hid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ====== API: 系统设置 ======
@app.route("/api/settings", methods=["GET"])
def api_settings():
    conn = get_db()
    rows = conn.execute("SELECT * FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in rows}
    conn.close()
    # 不返回密码
    settings.pop("password", None)
    return jsonify({"success": True, "settings": settings})


@app.route("/api/settings", methods=["POST"])
def api_settings_update():
    data = request.get_json(silent=True) or {}
    for key, value in data.items():
        if key == "password" and value:
            value = hashlib.md5(value.encode()).hexdigest()
        set_setting(key, value)

    # 同步配置到引擎
    if "dingtalk_webhook" in data:
        engine.dingtalk.configure(
            data["dingtalk_webhook"],
            data.get("dingtalk_enabled", "0") == "1"
        )

    add_log("INFO", "settings", "更新系统设置")
    return jsonify({"success": True})


# ====== API: NTP校时 ======
@app.route("/api/ntp/sync", methods=["POST"])
def api_ntp_sync():
    server = get_setting("ntp_server", "ntp.aliyun.com")
    success, result = engine.ntp.sync(server)
    if success:
        set_setting("ntp_last_sync", datetime.now().isoformat())
        return jsonify({"success": True, "offset": result})
    return jsonify({"success": False, "error": result}), 500


# ====== API: 钉钉测试 ======
@app.route("/api/dingtalk/test", methods=["POST"])
def api_dingtalk_test():
    webhook = get_setting("dingtalk_webhook", "")
    if not webhook:
        return jsonify({"error": "未配置钉钉Webhook"}), 400
    engine.dingtalk.configure(webhook, True)
    # 测试绕过冷却，直接发
    result = engine.dingtalk.send("测试告警", "这是一条测试消息，验证钉钉机器人配置是否正确。", bypass_cooldown=True)
    return jsonify({"success": result})


# ====== API: 音频设备列表 ======
@app.route("/api/audio_devices")
def api_audio_devices():
    devices = engine.player.get_audio_devices()
    return jsonify({"success": True, "devices": devices})


# ====== API: 运行日志 ======
@app.route("/api/logs")
def api_logs():
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    category = request.args.get("category", "")

    conn = get_db()
    if category:
        total = conn.execute(
            "SELECT COUNT(*) FROM logs WHERE category=?", (category,)
        ).fetchone()[0]
        rows = conn.execute(
            "SELECT * FROM logs WHERE category=? ORDER BY id DESC LIMIT ? OFFSET ?",
            (category, per_page, (page - 1) * per_page)
        ).fetchall()
    else:
        total = conn.execute("SELECT COUNT(*) FROM logs").fetchone()[0]
        rows = conn.execute(
            "SELECT * FROM logs ORDER BY id DESC LIMIT ? OFFSET ?",
            (per_page, (page - 1) * per_page)
        ).fetchall()

    conn.close()
    return jsonify({
        "success": True,
        "logs": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
    })


# ====== API: 课表导入导出 ======
@app.route("/api/export", methods=["GET"])
def api_export():
    """导出课表为Excel，支持?schedule_id=参数导出单个课表"""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "需要安装openpyxl: pip install openpyxl"}), 400

    conn = get_db()
    sid = request.args.get("schedule_id", "")
    
    if sid:
        schedules = conn.execute("SELECT * FROM schedules WHERE id=?", (sid,)).fetchall()
    else:
        schedules = conn.execute("SELECT * FROM schedules ORDER BY id").fetchall()

    wb = openpyxl.Workbook()
    area_names = {"A": "教学楼", "B": "操场", "C": "宿舍楼"}
    
    for sched in schedules:
        ws = wb.create_sheet(title=sched["name"][:31])
        # 表头（含中文说明）
        ws.append(["区域(A/B/C)", "时间(HH:MM)", "铃声类型", "任务名称", "适用日期(1-5=工作日)", "一次性日期(留空=重复)"])
        ws.append(["A", "08:00", "class_start", "上课铃", "1,2,3,4,5", ""])
        ws.append(["A", "08:45", "class_end", "下课铃", "1,2,3,4,5", ""])
        # 空行隔开
        ws.append([])
        # 实际数据
        tasks = conn.execute(
            "SELECT * FROM schedule_tasks WHERE schedule_id=? ORDER BY zone_id, time",
            (sched["id"],)
        ).fetchall()
        for t in tasks:
            ws.append([
                t["zone_id"], t["time"], t["bell_type"],
                t["task_name"], t["days"], t["one_time_date"] or ""
            ])

    # 删除默认sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    conn.close()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"课表_{date.today().isoformat()}.xlsx"
    )


@app.route("/api/template", methods=["GET"])
def api_template():
    """下载课表Excel模板"""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "需要安装openpyxl: pip install openpyxl"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课表模板"
    
    # 使用说明
    ws.append(["校园广播系统 - 课表导入模板"])
    ws.append(["填写说明：区域填 A/B/C/ALL，ALL表示全区域；时间格式 HH:MM；铃声类型见下方对照表；适用日期填数字逗号分隔"])
    ws.append(["铃声类型对照：class_start=上课铃 class_end=下课铃 prepare=预备铃 exercise=课间操 lunch=午餐铃 school_end=放学铃 emergency=紧急铃 custom=自定义"])
    ws.append([])
    ws.append(["区域", "时间", "铃声类型", "任务名称", "适用日期", "一次性日期"])
    # 示例数据
    examples = [
        ["A", "08:00", "class_start", "第一节课", "1,2,3,4,5", ""],
        ["A", "08:45", "class_end", "第一节课下课", "1,2,3,4,5", ""],
        ["A", "08:55", "prepare", "预备铃", "1,2,3,4,5", ""],
        ["A", "09:00", "class_start", "第二节课", "1,2,3,4,5", ""],
        ["A", "09:45", "class_end", "第二节课下课", "1,2,3,4,5", ""],
        ["ALL", "10:00", "exercise", "课间操", "1,2,3,4,5", ""],
        ["B", "10:05", "exercise", "课间操-操场", "1,2,3,4,5", ""],
        ["A", "12:00", "lunch", "午餐铃", "1,2,3,4,5", ""],
        ["A", "16:30", "school_end", "放学铃", "1,2,3,4,5", ""],
        ["C", "06:30", "wake_up", "起床铃", "1,2,3,4,5", ""],
        ["C", "22:00", "bedtime", "就寝铃", "1,2,3,4,5", ""],
    ]
    for ex in examples:
        ws.append(ex)
    
    # 删除默认sheet
    if "Sheet" in wb.sheetnames and ws.title != "Sheet":
        del wb["Sheet"]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="课表模板.xlsx"
    )


@app.route("/api/import", methods=["POST"])
def api_import():
    """导入课表"""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "需要安装openpyxl: pip install openpyxl"}), 400

    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400

    f = request.files["file"]
    schedule_name = request.form.get("schedule_name", f.filename)

    wb = openpyxl.load_workbook(f)
    conn = get_db()

    imported = 0
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sched_name = ws_name if len(wb.sheetnames) > 1 else schedule_name

        # 查找或创建课表
        existing = conn.execute("SELECT * FROM schedules WHERE name=?", (sched_name,)).fetchone()
        if existing:
            sid = existing["id"]
            conn.execute("DELETE FROM schedule_tasks WHERE schedule_id=?", (sid,))
        else:
            conn.execute("INSERT INTO schedules (name, is_active) VALUES (?,0)", (sched_name,))
            sid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            zone_id = str(row[0]).strip().upper()
            time_str = str(row[1]).strip()
            bell_type = str(row[2]).strip() if len(row) > 2 and row[2] else "class_start"
            task_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            days = str(row[4]).strip() if len(row) > 4 and row[4] else "1,2,3,4,5"
            one_time = str(row[5]).strip() if len(row) > 5 and row[5] else ""

            conn.execute(
                "INSERT INTO schedule_tasks (schedule_id, zone_id, time, bell_type, task_name, days, one_time_date) VALUES (?,?,?,?,?,?,?)",
                (sid, zone_id, time_str, bell_type, task_name, days, one_time)
            )
            imported += 1

    conn.commit()
    conn.close()

    add_log("INFO", "import", f"导入课表: {imported}条任务")
    return jsonify({"success": True, "imported": imported})


# ====== API: 健康检查 ======
@app.route("/favicon.ico")
def favicon():
    return "", 204

@app.route("/api/health")
def api_health():
    return jsonify({"status": "ok", "time": datetime.now().isoformat()})

# ====== API: 预约控制（bell_schedules） ======
@app.route("/api/bell_schedules", methods=["GET"])
def api_bell_schedules_list():
    """获取预约控制列表，支持?date=YYYY-MM-DD过滤"""
    target_date = request.args.get("date", "")
    conn = get_db()
    if target_date:
        rows = conn.execute(
            "SELECT * FROM bell_schedules WHERE date=? ORDER BY date, zone_id",
            (target_date,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM bell_schedules WHERE date >= date('now') ORDER BY date, zone_id"
        ).fetchall()
    conn.close()
    return jsonify({"success": True, "schedules": [dict(r) for r in rows]})


@app.route("/api/bell_schedules", methods=["POST"])
def api_bell_schedule_create():
    """创建预约控制"""
    data = request.get_json(silent=True) or {}
    target_date = data.get("date", "")
    zone_id = data.get("zone_id", "")
    action = data.get("action", "force_no_ring")
    schedule_id = data.get("schedule_id", 0)
    reason = data.get("reason", "")
    remind_enabled = data.get("remind_enabled", 1)

    if not target_date or not zone_id:
        return jsonify({"error": "日期和区域不能为空"}), 400
    if action not in ("force_ring", "force_no_ring"):
        return jsonify({"error": "操作类型无效，请使用 force_ring 或 force_no_ring"}), 400

    conn = get_db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO bell_schedules (date, zone_id, action, schedule_id, reason, remind_enabled, remind_at, reminded, created_at) VALUES (?,?,?,?,?,?,?,?,?)",
            (target_date, zone_id, action, schedule_id or 0, reason,
             1 if remind_enabled else 0, "20:00", 0,
             datetime.now().isoformat())
        )
        conn.commit()
        action_text = "打铃" if action == "force_ring" else "不打铃"
        add_log("INFO", "schedule", f"预约控制: {target_date} {zone_id} {action_text}")
        return jsonify({"success": True, "message": f"已预约 {target_date} {zone_id} {action_text}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/bell_schedules/<int:sid>", methods=["PUT"])
def api_bell_schedule_update(sid):
    """更新预约控制"""
    data = request.get_json(silent=True) or {}
    conn = get_db()
    existing = conn.execute("SELECT * FROM bell_schedules WHERE id=?", (sid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "预约不存在"}), 404

    fields = []
    values = []
    for key in ["date", "zone_id", "action", "schedule_id", "reason", "remind_enabled"]:
        if key in data:
            val = data[key]
            if key == "remind_enabled":
                val = 1 if val else 0
            fields.append(f"{key}=?")
            values.append(val)
    if fields:
        values.append(sid)
        conn.execute(f"UPDATE bell_schedules SET {','.join(fields)} WHERE id=?", values)
        conn.commit()
        add_log("INFO", "schedule", f"更新预约控制 ID={sid}")
    conn.close()
    return jsonify({"success": True})


@app.route("/api/bell_schedules/<int:sid>", methods=["DELETE"])
def api_bell_schedule_delete(sid):
    """删除预约控制"""
    conn = get_db()
    conn.execute("DELETE FROM bell_schedules WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    add_log("INFO", "schedule", f"删除预约控制 ID={sid}")
    return jsonify({"success": True})
